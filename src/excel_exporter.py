"""
excel_exporter.py
─────────────────────────────────────────────────────────────────────────────
LibFile 데이터 구조를 Excel (.xlsx) 파일로 내보내는 모듈입니다.

생성되는 시트:
    1. Matrix View (행렬 뷰)
         - 가로축(열): 모델명
         - 세로축(행): 파라미터명
         - 각 LIB 블록별로 행렬이 반복됩니다.
         - 어느 모델에 없는 파라미터는 빈 칸으로 처리합니다.

    2. List View (리스트 뷰)
         - Library / Model / Type / Parameter / Value 의 1차원 표 구조
         - 전체 데이터를 순서대로 한 줄씩 나열합니다.

의존성:
    - openpyxl: pip install openpyxl
─────────────────────────────────────────────────────────────────────────────
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from data_model import LibFile


def export_lib_to_excel(lib_file: LibFile, default_path: str = "lib_export.xlsx") -> str:
    """
    LibFile 객체의 데이터를 Excel 파일(.xlsx)로 내보내는 함수입니다.

    시트 구성:
        ┌─────────────────────────────────────────────────┐
        │ [Sheet 1] Matrix View                           │
        │  LIB: NMOS_TT                                   │
        │  Parameter │ NMOS_V1 │ NMOS_V2 │ NMOS_V3 │ …   │
        │  VTH0      │ 0.45    │ 0.47    │ 0.43    │ …   │
        │  TOX       │ 1.2e-8  │ 1.2e-8  │         │ …   │
        ├─────────────────────────────────────────────────┤
        │ [Sheet 2] List View                             │
        │  Library │ Model   │ Type │ Parameter │ Value   │
        │  NMOS_TT │ NMOS_V1 │ NMOS │ VTH0      │ 0.45   │
        │  NMOS_TT │ NMOS_V1 │ NMOS │ TOX       │ 1.2e-8 │
        └─────────────────────────────────────────────────┘

    처리 흐름:
        1. openpyxl Workbook / Worksheet 생성
        2. 각 LIB 블록별로 Matrix View 행렬 작성
        3. List View 시트에 전체 데이터를 1행씩 기록
        4. 모든 시트의 열 너비를 셀 내용 기준으로 자동 조정
        5. 지정 경로에 .xlsx 파일로 저장

    Args:
        lib_file     (LibFile): 파싱된 Smart Spice LIB 파일 데이터 객체
        default_path (str)    : 저장할 Excel 파일의 경로 (기본값: "lib_export.xlsx")

    Returns:
        str: 실제 저장된 파일의 경로
    """
    wb = openpyxl.Workbook()  # 새 워크북(Excel 파일) 생성

    # ─────────────────────────────────────────────────────────────────────
    # ── 시트 1: 행렬 뷰 (Matrix View) ──
    # 각 LIB 블록마다 [LIB 이름 헤더 → 모델명 행 → 파라미터별 값 행]을 반복합니다.
    # ─────────────────────────────────────────────────────────────────────
    ws_matrix = wb.active          # 첫 번째 시트를 Matrix View로 사용
    ws_matrix.title = "Matrix View"

    current_row = 1                # 현재 작성 중인 행 번호 (1-indexed)

    # 스타일 정의
    header_font  = Font(bold=True)                              # 헤더 셀용 굵은 글꼴
    center_align = Alignment(horizontal="center", vertical="center")  # 가운데 정렬

    for lib_block in lib_file.lib_blocks:
        # 모델이 없는 LIB 블록은 건너뜀
        if not lib_block.models:
            continue

        # ── LIB 블록 제목 행 출력 ──
        # 예: "LIB: NMOS_TT" 를 bold + 큰 폰트로 표시
        cell = ws_matrix.cell(row=current_row, column=1, value=f"LIB: {lib_block.name}")
        cell.font = Font(bold=True, size=12)
        current_row += 1

        models = lib_block.models  # 이 LIB 블록의 모델 목록

        # ── 모델명 열 헤더 행 ──
        # 1열: "Parameter" 라벨, 2열부터 각 모델명을 순서대로 기입
        ws_matrix.cell(row=current_row, column=1, value="Parameter").font = header_font
        for col_idx, model in enumerate(models, start=2):
            c = ws_matrix.cell(row=current_row, column=col_idx, value=model.name)
            c.font      = header_font
            c.alignment = center_align
        current_row += 1

        # ── 이 LIB 블록에 등장하는 모든 파라미터 이름 수집 ──
        # 각 모델마다 다른 파라미터를 가질 수 있으므로,
        # 중복 없이 선언 순서를 최대한 유지하면서 파라미터명 목록을 만듭니다.
        param_names = []
        for model in models:
            for p_name in model.params.keys():
                if p_name not in param_names:
                    param_names.append(p_name)  # 처음 등장한 파라미터명만 추가

        # ── 파라미터별 값 행 작성 ──
        # 각 파라미터에 대해 모든 모델의 값을 해당 열에 기입합니다.
        # 해당 모델에 그 파라미터가 없으면 빈 문자열("")로 처리합니다.
        for p_name in param_names:
            ws_matrix.cell(row=current_row, column=1, value=p_name)  # 1열: 파라미터명
            for col_idx, model in enumerate(models, start=2):
                val = model.params.get(p_name, "")  # 없으면 빈 칸
                ws_matrix.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 2  # 다음 LIB 블록과 구분하는 빈 줄 2칸

    # ─────────────────────────────────────────────────────────────────────
    # ── 시트 2: 리스트 뷰 (List View) ──
    # Library / Model / Type / Parameter / Value 5개 열로 구성됩니다.
    # ─────────────────────────────────────────────────────────────────────
    ws_list = wb.create_sheet(title="List View")

    # 헤더 행 작성
    headers = ["Library", "Model", "Type", "Parameter", "Value"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws_list.cell(row=1, column=col_idx, value=h)
        c.font = header_font   # 헤더는 굵은 글꼴

    list_row = 2  # 헤더 다음 줄부터 데이터 기록

    # 전체 LIB 블록 → 모델 → 파라미터 순으로 순회하며 1행씩 기록
    for lib_block in lib_file.lib_blocks:
        for model in lib_block.models:
            for p_name, p_val in model.params.items():
                ws_list.cell(row=list_row, column=1, value=lib_block.name)  # Library
                ws_list.cell(row=list_row, column=2, value=model.name)       # Model
                ws_list.cell(row=list_row, column=3, value=model.model_type) # Type
                ws_list.cell(row=list_row, column=4, value=p_name)           # Parameter
                ws_list.cell(row=list_row, column=5, value=p_val)            # Value
                list_row += 1

    # ─────────────────────────────────────────────────────────────────────
    # ── 열 너비 자동 조정 ──
    # 각 시트의 모든 열에 대해 셀 내용 중 가장 긴 문자열 길이를 측정하고
    # 그에 맞춰 열 너비를 설정합니다 (최소 여유 2칸 추가).
    # ─────────────────────────────────────────────────────────────────────
    for ws in [ws_matrix, ws_list]:
        for col in ws.columns:
            max_length  = 0
            col_letter  = col[0].column_letter  # 열 문자 (A, B, C, ...)
            for cell in col:
                try:
                    # 셀 값을 문자열로 변환하여 길이 측정
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass  # 값을 읽을 수 없는 셀은 건너뜀
            adjusted_width = (max_length + 2)  # 여유 2칸 추가
            ws.column_dimensions[col_letter].width = adjusted_width

    # ── 파일 저장 ──
    wb.save(default_path)
    return default_path
